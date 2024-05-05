from openpyxl import Workbook
from openpyxl import load_workbook
import os



cell_dict = {


    "A": 1,
    "B": 2,
    "C": 3,
    "D": 4,
    "E": 5,
    "F": 6,
    "G": 7,
    "H": 8,
    "I": 9,
    "J": 10,
    "K": 11,
    "L": 12,
    "M": 13,
    "N": 14,
    "O": 15,
    "P": 16,
    "Q": 17,
    "R": 18,
    "S": 19,
    "T": 20,
    "U": 21,
    "V": 22,
    "W": 23,
    "X": 24,
    "Y": 25,
    "Z": 26,


}


#populate colums/rows
#sheet is excel sheet


def check_sheet_exsits():
    wb = None
    ws = None
    sheet = input("Enter excel sheet name")
    sheet_path = sheet+".xlsx"
    cwd = os.getcwd()


    if sheet_path not in os.listdir( cwd ):
        prompt = input("Do you want to create file?").lower()
        if prompt == 'y':
            file_name = sheet_path
            wb = Workbook()
            wb.save(file_name)
            ws = wb.active


        else:
            return False, False, False
    else:
        wb = load_workbook( sheet_path )
        ws = wb.active
       
    return ws, wb, sheet_path


def populate():


    worksheet, workbook, sheet_path = check_sheet_exsits()
    if worksheet == False:
        return
   
    #input data / .txt file
    file_name = input("Enter text File Name")+".txt"


    if file_name not in os.listdir( os.getcwd() ):
        return False
   
    fhand = open(file_name, "r")


    #row/col user wants to populate
    row_col_start = input("Enter Col,Row Start Position").split(",")
    print(row_col_start)
    col_start = cell_dict[row_col_start[0]]
    row_start = int(row_col_start[1])




    row_col_end = input("Enter Row,Col End Position").split(",")
    col_end = cell_dict[row_col_end[0]]
    row_end = int(row_col_end[1])


    counter = 1
    big_data = []


    for row in fhand:
       
        #if we want to populate 3 columns 2 rows
        #but we have 5 rows of data, that wont work
        #columns start indexing at 1
        if counter-1 > col_end - col_start:
            return False
       
        little_data = row.split()
        print(little_data)


        #similar idea, if we want to do 2 columns
        #but we have 3 data pieces - dont want that bruh
        #row indexing starts at 1 too
        if len(little_data)-1 > row_end - row_start:
            print("Data doesnt fit")
            return False
       
        big_data.append(little_data)
       
        counter +=1
   
    i = 0
    for cols in worksheet.iter_rows(min_row = row_start, min_col = col_start, max_col = col_end, max_row = row_end ):
        cur_data = big_data[i]
        i+=1
        for cell, data in zip(cols, cur_data):
            cell.value = data


       
    workbook.save(sheet_path)
    print("Done")


def delete():
   


    worksheet, workbook, sheet_path = check_sheet_exsits()
    if worksheet == False:
        return


    var = input('Would you like to delete a cell, row, or coloumn? ')


    if var == 'cell':
       cell = input('What cell would you like to delete?')
       worksheet[cell] = ""
 
    elif var == "row":
       inp_row = int(input('What row would you like to delete?'))
       for row in worksheet.iter_rows(max_row = inp_row):
           if row[0].row == inp_row:
               for cell in row:
                   cell.value = ""




    elif var == "col":
        inp_col = int(input('What row would you like to delete?'))
        for column in worksheet.iter_cols(max_row = inp_col):
            if column[0].col == inp_col:
                for cell in column:
                   cell.value = ""
 
    workbook.save(sheet_path)


 



